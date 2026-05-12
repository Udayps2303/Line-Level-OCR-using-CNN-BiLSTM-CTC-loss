import os
from collections import defaultdict
from statistics import median, multimode

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import mplcursors
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Path to mapping file ─────────────────────────────────────────────────────
mapping_file = "output/map_sorted.txt"

# ── Stats containers ──────────────────────────────────────────────────────────
total_lines = 0
total_words = 0
total_chars = 0

lines_per_book           = defaultdict(int)
words_per_book           = defaultdict(int)
chars_per_book           = defaultdict(int)

words_per_line_list      = []
chars_per_line_list      = []

words_per_line_per_book  = defaultdict(list)
chars_per_line_per_book  = defaultdict(list)

with open(mapping_file, "r", encoding="utf-8") as f:
    for line in f:
        parts = line.strip().split("\t")
        if len(parts) != 2:
            continue
        img_path, text = parts
        book       = img_path.split("/")[1]
        words      = text.strip().split()
        word_count = len(words)

        # Character count: count Unicode characters in the text (excluding
        # leading/trailing whitespace). For Odia script each akshara is one
        # Unicode code point, so len() gives the correct character count.
        char_count = len(text.strip())

        total_lines += 1
        total_words += word_count
        total_chars += char_count

        lines_per_book[book]          += 1
        words_per_book[book]          += word_count
        chars_per_book[book]          += char_count

        words_per_line_list.append(word_count)
        chars_per_line_list.append(char_count)

        words_per_line_per_book[book].append(word_count)
        chars_per_line_per_book[book].append(char_count)

# ── Helper ────────────────────────────────────────────────────────────────────
def book_sort_key(book_name):
    part = book_name.split("_")[0]
    return int(part) if part.isdigit() else float("inf")

avg_words_per_line = total_words / total_lines if total_lines else 0
overall_median_wpl = median(words_per_line_list)
overall_modes_wpl  = multimode(words_per_line_list)

avg_chars_per_line = total_chars / total_lines if total_lines else 0
overall_median_cpl = median(chars_per_line_list)
overall_modes_cpl  = multimode(chars_per_line_list)

# ── Print stats ───────────────────────────────────────────────────────────────
print("\n===== OVERALL STATS =====")
print(f"Total lines:               {total_lines}")
print(f"Total words:               {total_words}")
print(f"Total characters:          {total_chars}")
print(f"Average words per line:    {avg_words_per_line:.2f}")
print(f"Median words per line:     {overall_median_wpl:.2f}")
print(f"Mode words per line:       {overall_modes_wpl[0] if len(overall_modes_wpl) == 1 else overall_modes_wpl}")
print(f"Average chars per line:    {avg_chars_per_line:.2f}")
print(f"Median chars per line:     {overall_median_cpl:.2f}")
print(f"Mode chars per line:       {overall_modes_cpl[0] if len(overall_modes_cpl) == 1 else overall_modes_cpl}")

print("\n===== PER BOOK STATS =====")
for book in sorted(lines_per_book.keys(), key=book_sort_key):
    lines  = lines_per_book[book]
    words  = words_per_book[book]
    chars  = chars_per_book[book]
    wpl    = words_per_line_per_book[book]
    cpl    = chars_per_line_per_book[book]

    avg_w  = words / lines if lines else 0
    med_w  = median(wpl)
    modes_w = multimode(wpl)
    mod_w_str = str(modes_w[0]) if len(modes_w) == 1 else str(modes_w)

    avg_c  = chars / lines if lines else 0
    med_c  = median(cpl)
    modes_c = multimode(cpl)
    mod_c_str = str(modes_c[0]) if len(modes_c) == 1 else str(modes_c)

    print(f"\nBook: {book}")
    print(f"  Lines:                  {lines}")
    print(f"  Words:                  {words}")
    print(f"  Characters:             {chars}")
    print(f"  Avg  words/line:        {avg_w:.2f}")
    print(f"  Median words/line:      {med_w:.2f}")
    print(f"  Mode words/line:        {mod_w_str}")
    print(f"  Avg  chars/line:        {avg_c:.2f}")
    print(f"  Median chars/line:      {med_c:.2f}")
    print(f"  Mode chars/line:        {mod_c_str}")

# ── Figure 1: Word distribution histogram ────────────────────────────────────
fig1, ax1 = plt.subplots(figsize=(11, 6))
n1, bins1, patches1 = ax1.hist(words_per_line_list, bins=25, color="steelblue", edgecolor="white")

for count, patch in zip(n1, patches1):
    if count > 0:
        ax1.text(
            patch.get_x() + patch.get_width() / 2,
            patch.get_height(),
            f"{int(count)}",
            ha="center", va="bottom",
            fontsize=7.5, fontweight="bold", color="#222222",
        )

ax1.set_xlabel("Number of Words per Line", fontsize=12)
ax1.set_ylabel("Number of Lines",          fontsize=12)
ax1.set_title("Distribution of Words per Line", fontsize=14, fontweight="bold")
ax1.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
ax1.grid(axis="y", alpha=0.4)

cursor1 = mplcursors.cursor(patches1, hover=True)

@cursor1.connect("add")
def on_add_words(sel):
    patch = sel.artist
    x, w, y = patch.get_x(), patch.get_width(), patch.get_height()
    sel.annotation.set_text(f"Words: {x:.1f} – {x + w:.1f}\nCount: {int(y)}")
    sel.annotation.get_bbox_patch().set(fc="lightyellow", alpha=0.9)

plt.tight_layout()

# ── Figure 2: Character distribution histogram ───────────────────────────────
fig2, ax2 = plt.subplots(figsize=(11, 6))
n2, bins2, patches2 = ax2.hist(chars_per_line_list, bins=25, color="darkorange", edgecolor="white")

for count, patch in zip(n2, patches2):
    if count > 0:
        ax2.text(
            patch.get_x() + patch.get_width() / 2,
            patch.get_height(),
            f"{int(count)}",
            ha="center", va="bottom",
            fontsize=7.5, fontweight="bold", color="#222222",
        )

ax2.set_xlabel("Number of Characters per Line", fontsize=12)
ax2.set_ylabel("Number of Lines",               fontsize=12)
ax2.set_title("Distribution of Characters per Line", fontsize=14, fontweight="bold")
ax2.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
ax2.grid(axis="y", alpha=0.4)

cursor2 = mplcursors.cursor(patches2, hover=True)

@cursor2.connect("add")
def on_add_chars(sel):
    patch = sel.artist
    x, w, y = patch.get_x(), patch.get_width(), patch.get_height()
    sel.annotation.set_text(f"Chars: {x:.1f} – {x + w:.1f}\nCount: {int(y)}")
    sel.annotation.get_bbox_patch().set(fc="lightyellow", alpha=0.9)

plt.tight_layout()
plt.show()

# ── Excel export ──────────────────────────────────────────────────────────────
OUTPUT_EXCEL = "output/stats_report.xlsx"
os.makedirs("output", exist_ok=True)

wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL   = PatternFill("solid", start_color="2E4057")
TITLE_FONT    = Font(name="Arial", bold=True, size=13, color="2E4057")
LABEL_FONT    = Font(name="Arial", bold=True, size=10)
VALUE_FONT    = Font(name="Arial", size=10)
SUBHEAD_FILL  = PatternFill("solid", start_color="D9E1F2")
SUBHEAD_FONT  = Font(name="Arial", bold=True, size=10, color="2E4057")
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
RIGHT         = Alignment(horizontal="right",  vertical="center")
THIN          = Side(style="thin", color="AAAAAA")
THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL      = PatternFill("solid", start_color="EEF2FA")

# Orange-tinted header for character sheets
CHAR_HEADER_FILL = PatternFill("solid", start_color="7B3F00")

def style_header_cell(cell, text, fill=None):
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = fill if fill else HEADER_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER

def style_value_cell(cell, value, align=RIGHT):
    cell.value     = value
    cell.font      = VALUE_FONT
    cell.alignment = align
    cell.border    = THIN_BORDER

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Overall Stats  (words + characters combined)
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Overall Stats"

ws1.merge_cells("A1:B1")
title_cell        = ws1["A1"]
title_cell.value  = "Overall Dataset Statistics"
title_cell.font   = TITLE_FONT
title_cell.alignment = CENTER
ws1.row_dimensions[1].height = 28

overall_rows = [
    ("Total Lines",                total_lines),
    ("Total Words",                total_words),
    ("Total Characters",           total_chars),
    ("Average Words per Line",     round(avg_words_per_line, 2)),
    ("Median Words per Line",      round(float(overall_median_wpl), 2)),
    ("Mode Words per Line",        overall_modes_wpl[0] if len(overall_modes_wpl) == 1 else str(overall_modes_wpl)),
    ("Average Characters per Line",round(avg_chars_per_line, 2)),
    ("Median Characters per Line", round(float(overall_median_cpl), 2)),
    ("Mode Characters per Line",   overall_modes_cpl[0] if len(overall_modes_cpl) == 1 else str(overall_modes_cpl)),
]

for i, (label, value) in enumerate(overall_rows, start=2):
    lc = ws1.cell(row=i, column=1, value=label)
    lc.font      = LABEL_FONT
    lc.alignment = LEFT
    lc.border    = THIN_BORDER
    lc.fill      = SUBHEAD_FILL if i % 2 == 0 else PatternFill()

    vc = ws1.cell(row=i, column=2, value=value)
    vc.font      = VALUE_FONT
    vc.alignment = RIGHT
    vc.border    = THIN_BORDER
    if i % 2 == 0:
        vc.fill = SUBHEAD_FILL

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 22

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Per-Book Stats (words)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Per-Book Stats (Words)")

headers_w = ["Book", "Lines", "Words", "Avg Words/Line", "Median Words/Line", "Mode Words/Line"]
for col, h in enumerate(headers_w, 1):
    style_header_cell(ws2.cell(row=1, column=col), h)
ws2.row_dimensions[1].height = 22

for row_idx, book in enumerate(sorted(lines_per_book.keys(), key=book_sort_key), start=2):
    lines  = lines_per_book[book]
    words  = words_per_book[book]
    wpl    = words_per_line_per_book[book]
    avg    = round(words / lines, 2) if lines else 0
    med    = round(float(median(wpl)), 2)
    modes  = multimode(wpl)
    mod_val = modes[0] if len(modes) == 1 else str(modes)

    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    row_data = [book, lines, words, avg, med, mod_val]
    aligns   = [LEFT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT]

    for col_idx, (val, aln) in enumerate(zip(row_data, aligns), 1):
        cell           = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = VALUE_FONT
        cell.alignment = aln
        cell.border    = THIN_BORDER
        cell.fill      = fill

col_widths_w = [14, 10, 12, 18, 20, 22]
for col_idx, width in enumerate(col_widths_w, 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Per-Book Stats (characters)
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Per-Book Stats (Chars)")

headers_c = ["Book", "Lines", "Characters", "Avg Chars/Line", "Median Chars/Line", "Mode Chars/Line"]
for col, h in enumerate(headers_c, 1):
    style_header_cell(ws3.cell(row=1, column=col), h, fill=CHAR_HEADER_FILL)
ws3.row_dimensions[1].height = 22

for row_idx, book in enumerate(sorted(lines_per_book.keys(), key=book_sort_key), start=2):
    lines  = lines_per_book[book]
    chars  = chars_per_book[book]
    cpl    = chars_per_line_per_book[book]
    avg    = round(chars / lines, 2) if lines else 0
    med    = round(float(median(cpl)), 2)
    modes  = multimode(cpl)
    mod_val = modes[0] if len(modes) == 1 else str(modes)

    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    row_data = [book, lines, chars, avg, med, mod_val]
    aligns   = [LEFT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT]

    for col_idx, (val, aln) in enumerate(zip(row_data, aligns), 1):
        cell           = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = VALUE_FONT
        cell.alignment = aln
        cell.border    = THIN_BORDER
        cell.fill      = fill

col_widths_c = [14, 10, 14, 18, 20, 22]
for col_idx, width in enumerate(col_widths_c, 1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = width

ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Histogram Data (words)
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Histogram Data (Words)")
hist_headers_w = ["Bin Start", "Bin End", "Line Count"]
for col, h in enumerate(hist_headers_w, 1):
    style_header_cell(ws4.cell(row=1, column=col), h)

counts_w, bin_edges_w = np.histogram(words_per_line_list, bins=25)
for row_idx, (cnt, b_start, b_end) in enumerate(
        zip(counts_w, bin_edges_w[:-1], bin_edges_w[1:]), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for col_idx, val in enumerate([round(float(b_start), 2), round(float(b_end), 2), int(cnt)], 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = VALUE_FONT
        cell.alignment = RIGHT
        cell.border    = THIN_BORDER
        cell.fill      = fill

for col_idx in range(1, 4):
    ws4.column_dimensions[get_column_letter(col_idx)].width = 16

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Histogram Data (characters)
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Histogram Data (Chars)")
hist_headers_c = ["Bin Start", "Bin End", "Line Count"]
for col, h in enumerate(hist_headers_c, 1):
    style_header_cell(ws5.cell(row=1, column=col), h, fill=CHAR_HEADER_FILL)

counts_c, bin_edges_c = np.histogram(chars_per_line_list, bins=25)
for row_idx, (cnt, b_start, b_end) in enumerate(
        zip(counts_c, bin_edges_c[:-1], bin_edges_c[1:]), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for col_idx, val in enumerate([round(float(b_start), 2), round(float(b_end), 2), int(cnt)], 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = VALUE_FONT
        cell.alignment = RIGHT
        cell.border    = THIN_BORDER
        cell.fill      = fill

for col_idx in range(1, 4):
    ws5.column_dimensions[get_column_letter(col_idx)].width = 16

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_EXCEL)
print(f"\n✅  Excel report saved → {OUTPUT_EXCEL}")